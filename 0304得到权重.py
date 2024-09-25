# 导入所需的库
import pandas as pd
import numpy as np
from scipy.stats import entropy
from sklearn.preprocessing import StandardScaler, MinMaxScaler
from openpyxl import load_workbook, Workbook

# 定义输出文件路径
output_file_path = r"J:\BaiduNetdiskDownload\05马克\8亿+企业查数据\03指标计算\企业查区县统计合并_新指标和原始指标merged_data归一化.xlsx"


# 定义TopsisEntropy类，用于实现基于熵权法的TOPSIS评价方法
class TopsisEntropy:
    def __init__(self, indicators, alternatives, data):
        self.indicators = indicators  # 评价指标
        self.alternatives = alternatives  # 评价对象
        self.data = self.preprocess_data(data)

    def preprocess_data(self, data):
        processed_data = []
        for i in range(data.shape[1]):
            column = data[:, i]
            try:
                # 尝试将列转换为浮点数
                processed_column = pd.to_numeric(column, errors='coerce')
                # 检查是否有 NaN 值
                if np.isnan(processed_column).any():
                    print(f"警告：列 '{self.indicators[i]}' 包含无法转换为数值的数据，这些值将被替换为 0")
                    processed_column = np.nan_to_num(processed_column, 0)
                processed_data.append(processed_column)
            except Exception as e:
                print(f"错误：处理列 '{self.indicators[i]}' 时出现问题：{str(e)}")
                raise
        return np.array(processed_data).T

    # 计算熵权
    def calculate_entropy_weights(self):
        n, m = self.data.shape
        entropy_list = []

        for i in range(m):
            column_data = self.data[:, i]

            # 处理零值和极小值
            column_data = np.clip(column_data, 1e-10, None)
            column_sum = np.sum(column_data)

            # 如果列和为零，则将权重设为零
            if column_sum == 0:
                entropy_list.append(0)
                continue

            p = column_data / column_sum
            e = -np.sum(p * np.log(p + 1e-10)) / np.log(n)
            entropy_list.append(e)

        # 计算权重
        weight_list = [1 - e for e in entropy_list]
        weight_sum = np.sum(weight_list)

        # 如果所有权重都为零，则平均分配
        if weight_sum == 0:
            self.weight_list = np.ones(m) / m
        else:
            self.weight_list = np.array(weight_list) / weight_sum

    # 数据标准化
    def normalize_matrix(self):
        if np.any(self.data < 0):
            scaler = MinMaxScaler()
        else:
            scaler = StandardScaler()
        self.normalized_matrix = scaler.fit_transform(self.data)

    # 确定理想解和负理想解
    def determine_ideal_solution(self, maximize):
        if maximize:
            self.ideal_solution = np.max(self.normalized_matrix, axis=0)
            self.worst_solution = np.min(self.normalized_matrix, axis=0)
        else:
            self.ideal_solution = np.min(self.normalized_matrix, axis=0)
            self.worst_solution = np.max(self.normalized_matrix, axis=0)

    # 计算到理想解和负理想解的距离
    def calculate_distance(self):
        diff_ideal = self.ideal_solution - self.normalized_matrix
        diff_worst = self.worst_solution - self.normalized_matrix
        weight_matrix = np.reshape(self.weight_list, (1, -1))
        self.distance_to_ideal = np.sqrt(np.sum(weight_matrix * diff_ideal ** 2, axis=1))
        self.distance_to_worst = np.sqrt(np.sum(weight_matrix * diff_worst ** 2, axis=1))

    # 计算相对接近度
    def calculate_similarity(self):
        self.similarity = self.distance_to_worst / (self.distance_to_ideal + self.distance_to_worst)

    # TOPSIS评价过程
    def topsis(self, maximize=True):
        self.calculate_entropy_weights()
        self.normalize_matrix()
        self.determine_ideal_solution(maximize)
        self.calculate_distance()
        self.calculate_similarity()
        return self.similarity

    # 找出最佳方案
    def find_best_alternative(self):
        similarity = self.topsis(maximize=True)
        best_alternative = self.alternatives[np.argmax(similarity)]
        return best_alternative

    # 评估所有方案
    def evaluate_alternatives(self):
        self.calculate_entropy_weights()
        similarity = self.topsis(maximize=True)
        best_alternative = self.find_best_alternative()

        print("各个指标的信息熵权重为：", self.weight_list)
        print("每个方案的相似度为：", similarity)
        print("最佳方案是：", best_alternative)

        return self.weight_list, similarity, best_alternative

    # 导出结果到Excel文件
    def export_results(self, weight_list, similarity):
        # 确保 weight_list 和 indicators 长度相同
        if len(weight_list) != len(self.indicators):
            print(f"警告：权重列表长度 ({len(weight_list)}) 与指标数量 ({len(self.indicators)}) 不匹配")
            weight_list = weight_list[:len(self.indicators)]  # 截断权重列表以匹配指标数量

        # 导出指标的信息熵权重
        weights_df = pd.DataFrame({
            '三级指标': self.indicators,
            '信息熵权重': weight_list
        })

        # 确保 similarity 和 alternatives 长度相同
        if len(similarity) != len(self.alternatives):
            print(f"警告：相似度列表长度 ({len(similarity)}) 与方案数量 ({len(self.alternatives)}) 不匹配")
            similarity = similarity[:len(self.alternatives)]  # 截断相似度列表以匹配方案数量

        # 导出方案的相似度
        similarity_df = pd.DataFrame({
            '方案': self.alternatives,
            '相似度': similarity
        })

        # 检查文件是否存在
        try:
            book = load_workbook(output_file_path)
        except FileNotFoundError:
            book = Workbook()

        # 删除已存在的工作表（如果有）
        if '指标信息熵权重' in book.sheetnames:
            del book['指标信息熵权重']
        if '方案相似度' in book.sheetnames:
            del book['方案相似度']

        # 创建新的工作表并写入数据
        weights_sheet = book.create_sheet('指标信息熵权重')
        similarity_sheet = book.create_sheet('方案相似度')

        # 读取一二三级指标表
        indicator_levels_df = pd.read_excel(r"J:\BaiduNetdiskDownload\05马克\8亿+企业查数据\03指标计算\一二三级指标.xlsx")
        
        # 创建权重数据的DataFrame
        weights_df = pd.DataFrame({
            '三级指标': self.indicators,
            '信息熵权重': weight_list
        })
        
        # 与一二三级指标表进行join
        merged_weights_df = weights_df.merge(indicator_levels_df, on='三级指标', how='left')
        
        # 写入权重数据
        weights_sheet.append(['一级指标', '二级指标', '三级指标', '综合指数', '信息熵权重'])
        for _, row in merged_weights_df.iterrows():
            weights_sheet.append([row['一级指标'], row['二级指标'], row['三级指标'], row['综合指数'], row['信息熵权重']])

        # 写入相似度数据
        similarity_sheet.append(['方案', '相似度'])
        for alternative, sim in zip(self.alternatives, similarity):
            similarity_sheet.append([alternative, sim])

        # 保存文件
        book.save(output_file_path)

        print(f"结果已导出到文件：{output_file_path}")


# 读取数据
data = pd.read_excel(output_file_path)

# 获取指标和方案
indicators = data.columns[1:].tolist()  # 假设第一列是方案名称
print(indicators)
alternatives = data.iloc[:, 0].tolist()
#print(alternatives)

# 创建 TopsisEntropy 对象并进行评估
topsis = TopsisEntropy(indicators, alternatives, data.iloc[:, 1:].values)
weight_list, similarity, best_alternative = topsis.evaluate_alternatives()

# 导出结果
topsis.export_results(weight_list, similarity)

print(f"结果已导出到文件：{output_file_path}")
