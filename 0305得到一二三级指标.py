import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 读取数据
input_file_path = r'J:\BaiduNetdiskDownload\05马克\8亿+企业查数据\03指标计算\企业查区县统计合并_新指标和原始指标merged_data归一化.xlsx'
weights_file_path = r'J:\BaiduNetdiskDownload\05马克\8亿+企业查数据\03指标计算\企业查区县统计合并_新指标和原始指标merged_data归一化.xlsx'

df = pd.read_excel(input_file_path, sheet_name='Sheet1')
weights_df = pd.read_excel(weights_file_path, sheet_name='指标信息熵权重')

# 提取指标和相应的权重
weights = weights_df[['一级指标', '二级指标', '三级指标', '信息熵权重']]

# 计算二级指标
secondary_indicators = pd.DataFrame(index=df.index)
for secondary_indicator in weights['二级指标'].unique():
    related_weights = weights[weights['二级指标'] == secondary_indicator]
    weighted_sum = pd.Series(0, index=df.index)
    for _, row in related_weights.iterrows():
        tertiary_indicator = row['三级指标']
        weight = row['信息熵权重']
        if tertiary_indicator in df.columns:
            weighted_sum += df[tertiary_indicator] * weight
        else:
            print(f"警告：未找到指标 '{tertiary_indicator}' 的值，跳过此指标")
    secondary_indicators[secondary_indicator] = weighted_sum

# 计算一级指标
primary_indicators = pd.DataFrame(index=df.index)
for primary_indicator in weights['一级指标'].unique():
    related_weights = weights[weights['一级指标'] == primary_indicator]
    weighted_sum = pd.Series(0, index=df.index)
    for _, row in related_weights.iterrows():
        secondary_indicator = row['二级指标']
        weight = row['信息熵权重']
        if secondary_indicator in secondary_indicators.columns:
            weighted_sum += secondary_indicators[secondary_indicator] * weight
        else:
            print(f"警告：未找到指标 '{secondary_indicator}' 的值，跳过此指标")
    primary_indicators[primary_indicator] = weighted_sum

# 计算综合指数
comprehensive_index = primary_indicators.sum(axis=1)

# 合并所有结果
result_df = pd.concat([df, secondary_indicators, primary_indicators], axis=1)
result_df['综合指数'] = comprehensive_index

# 保存结果到文件中
output_file_path = input_file_path.replace('.xlsx', '_指标计算结果.xlsx')
result_df.to_excel(output_file_path, index=False)

# 加载保存好的excel文件并添加颜色
wb = load_workbook(output_file_path)
ws = wb.active

color_mapping = {
    '一级指标': 'FFFF00',  # 黄色
    '二级指标': '00FF00',  # 绿色
    '三级指标': '00BFFF',  # 蓝色
    '综合指数': 'FF0000'   # 红色
}

for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
    header = col[0].value
    if header == '综合指数':
        fill = PatternFill(start_color=color_mapping['综合指数'], end_color=color_mapping['综合指数'], fill_type="solid")
    else:
        for indicator_type, color in color_mapping.items():
            if header in weights_df[indicator_type].values:
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                break
        else:
            continue
    for cell in col:
        cell.fill = fill

# 保存带颜色的Excel表
wb.save(output_file_path)

print(f"指标计算完成并已保存到文件中：{output_file_path}")
print("同时为不同类型的指标赋予了不同的颜色")