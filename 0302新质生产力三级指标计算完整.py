import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, DataBarRule
from openpyxl.utils import get_column_letter

# 定义文件路径
base_folder = r"J:\BaiduNetdiskDownload\05马克\8亿+企业查数据"
input_excel_path = os.path.join(base_folder, "03指标计算", "企业查区县统计合并.xlsx")
boundary_path = os.path.join(base_folder, "01区县统计", "行政边界_gjdl_县.xlsx")
summary_path = os.path.join(base_folder, "03指标计算", "原始数据汇总df.xlsx")
output_excel_path = os.path.join(base_folder, "03指标计算", "新质生产力指标计算结果.xlsx")

# 读取数据
df = pd.read_excel(input_excel_path)
boundary_df = pd.read_excel(boundary_path)
summary_df = pd.read_excel(summary_path, sheet_name="数据汇总单")

# 合并数据
merged_df = boundary_df.merge(df, left_on="gjdl_xdm", right_on="district_code")
merged_df = merged_df.merge(summary_df, left_on="district_code", right_on="区县代码")

# # 修改异常数据
# merged_df.loc[merged_df['district_code'] == '440106', 'district_code'] = '440111'

# 修正特定区县的参保人数数据和GDP数据
special_districts = {
    '120114': {'扩展信息_insured_number_sum': 261300},
    '130109': {'扩展信息_insured_number_sum': 12481},
    '130110': {'扩展信息_insured_number_sum': 32191},
    '131103': {'扩展信息_insured_number_sum': 22152},
    '420204': {'扩展信息_insured_number_sum': 29428},
    '350505': {'A 2022区县GDP(亿元）': 696.59}
}
for district, values in special_districts.items():
    for column, value in values.items():
        merged_df.loc[merged_df['district_code'] == district, column] = value
for column in merged_df.columns:
    print(column)
# 计算新指标
new_indicators = {
    '科技人才比率': lambda df: df['专利信息_总和'] * 5 / df['扩展信息_insured_number_sum'],
    '高新技术产业从业人员占比': lambda df: df['高新技术企业_总和_参保人数'] / df['扩展信息_insured_number_sum'],
    '人均受教育年限': lambda df: df['K 人均受教育年限（年）'],
    '科研机构数量': lambda df: df['11poi_2023年数量'],
    '重大科技基础设施数量': lambda df: df['N 重大基础设施'],
    '区县R&D经费支出': lambda df: df['专利信息_发明专利']  ,#df['I 市级科学支出（2021）（万元）']
    '专利数量': lambda df: df['专利信息_总和'],
    '发明专利占比': lambda df: df['专利信息_发明专利'] / (df['专利信息_发明专利'] + df['专利信息_外观专利'] + df['专利信息_实用新型']),
    '有R&D活动的企业占比': lambda df: df['专利信息_专利信息_计数'] / df['工商注册基本信息_总数'],
    'GDP/人员': lambda df: df['A 2022区县GDP(亿元）'] * 100000000 / df['扩展信息_insured_number_sum'],
    '发明专利数量/全部人员数量': lambda df: df['专利信息_总和'] / df['扩展信息_insured_number_sum'],
    'GDP/实缴资本': lambda df: df['A 2022区县GDP(亿元）'] * 100000000 ,#/ df['工商注册基本信息_总和_real_capital']
    '专利专著证书/实缴资本': lambda df: (df['专利信息_专利信息_计数'] + df['作品著作权_总和'] + df['资质证书_unified_code_count']) / df['工商注册基本信息_总和_real_capital'],
    '绿色金融指数': lambda df: df['E 市绿色金融指数(2022)'],
    '单位GDP能耗': lambda df: df['C 单位GDP能耗(城市）'],
    '碳排放量/碳汇': lambda df: df['D 单位GDP碳排'],
    '新增新质生产力用地面积': lambda df: df['10土地出让_总计'],
    '绿色专利授权数量': lambda df: df['F 年均绿色专利申请数量（2016-2024）'],
    '专精特新小巨人企业专利应用': lambda df: df['O 专精特新小巨人专利利用情况'],
    '科研机构增长率': lambda df: df['11poi_差值(2023-2022)'] / df['11poi_2022年数量'],
    '产业结构高级化指数': lambda df: df['M 产业结构高级化指数'],
    '新增新质生产力相关企业数量': lambda df: df['高新技术企业_总和_计数'],
    '高新技术企业占比': lambda df: df['高新技术企业_总和_计数'],# / df['工商注册基本信息_总数']
    '居民收入or增速or收入支出比': lambda df: df['B 2022市级收入支出比'],
    '城乡收入差距': lambda df: df['P 城乡收入差距'],
    '建成区绿地覆盖率': lambda df: df['H 建成区绿地率'],
    '环境处罚': lambda df: df['行政处罚_环罚_unified_code_count'],
    '通过应用绿色技术的企业比例': lambda df: df['G 上市企业通过的绿色质量认证总数']
}

for indicator, formula in new_indicators.items():
    merged_df[indicator] = formula(merged_df)

# 定义要标记的列
blue_columns = list(new_indicators.keys())
red_columns = [
    '专利信息_总和', '扩展信息_insured_number_sum', '高新技术企业_总和_参保人数',
    'K 人均受教育年限（年）', '11poi_2023年数量', 'N 重大基础设施',
    'I 市级科学支出（2021）（万元）', '专利信息_发明专利', '专利信息_外观专利',
    '专利信息_实用新型', '专利信息_专利信息_计数', '工商注册基本信息_总数',
    'A 2022区县GDP(亿元）', '工商注册基本信息_总和_real_capital', '作品著作权_总和',
    '资质证书_unified_code_count', 'E 市绿色金融指数(2022)', 'C 单位GDP能耗(城市）',
    'D 单位GDP碳排', '10土地出让_总计', 'F 年均绿色专利申请数量（2016-2024）',
    'O 专精特新小巨人专利利用情况', '11poi_差值(2023-2022)', '11poi_2022年数量',
    'M 产业结构高级化指数', '高新技术企业_总和_计数', 'B 2022市级收入支出比',
    'P 城乡收入差距', 'H 建成区绿地率', '行政处罚_环罚_unified_code_count',
    'G 上市企业通过的绿色质量认证总数'
]

# 重新排列列顺序
other_columns = [col for col in merged_df.columns if col not in blue_columns and col not in red_columns]
new_column_order = other_columns + red_columns + blue_columns
merged_df = merged_df[new_column_order]

# 导出为Excel
with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
    # 创建一个新的DataFrame，包含district code列和blue_columns
    blue_df = merged_df[['district_code'] + blue_columns]

    # 将数据导出到Excel的Sheet1和Sheet2
    blue_df.to_excel(writer, index=False, sheet_name='Sheet1')
    merged_df.to_excel(writer, index=False, sheet_name='Sheet2')

    # 创建Sheet3并导出所需列
    sheet3_columns = ['shengshixian', 'district_code'] + blue_columns + red_columns
    sheet3_df = merged_df[sheet3_columns]
    sheet3_df.to_excel(writer, index=False, sheet_name='Sheet3')

    # 获取工作簿和工作表对象
    workbook = writer.book
    worksheet1 = writer.sheets['Sheet1']
    worksheet2 = writer.sheets['Sheet2']
    worksheet3 = writer.sheets['Sheet3']  # 添加这行来定义 worksheet3

    # 应用样式到Sheet1、Sheet2和Sheet3
    for worksheet, df in [(worksheet1, blue_df), (worksheet2, merged_df), (worksheet3, sheet3_df)]:
        for col_num, column in enumerate(df.columns, 1):
            if column in blue_columns or (column in red_columns and worksheet != worksheet1):
                # 为每一列设置条件格式数据条
                worksheet.conditional_formatting.add(
                    f'{get_column_letter(col_num)}2:{get_column_letter(col_num)}{worksheet.max_row}',
                    DataBarRule(
                        start_type='min',
                        end_type='max',
                        color='87CEFA' if column in blue_columns else 'FF0000',
                        showValue=True,
                        minLength=None,
                        maxLength=None
                    )
                )
                # 设置列标题的填充颜色
                worksheet.cell(row=1, column=col_num).fill = PatternFill(
                    start_color='87CEFA' if column in blue_columns else 'FF0000',
                    end_color='87CEFA' if column in blue_columns else 'FF0000',
                    fill_type='solid'
                )

print(f"新质生产力指标计算结果已导出为Excel: {output_excel_path}")