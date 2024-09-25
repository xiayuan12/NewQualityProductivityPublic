import pandas as pd
import numpy as np
from sklearn.preprocessing import MinMaxScaler, StandardScaler
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Color
from openpyxl.formatting.rule import ColorScaleRule

# 1. 读取Excel文件
file_path = r"J:\BaiduNetdiskDownload\05马克\8亿+企业查数据\03指标计算\新质生产力指标计算结果.xlsx"
df = pd.read_excel(file_path)

# 2. 查看数据
print(df.head())

# 3. 选择需要归一化的列
columns_to_normalize = [
    '科技人才比率', '高新技术产业从业人员占比', '人均受教育年限', '科研机构数量',
    '重大科技基础设施数量', '区县R&D经费支出', '专利数量', '发明专利占比',
    '有R&D活动的企业占比', 'GDP/人员', '发明专利数量/全部人员数量', 'GDP/实缴资本',
    '专利专著证书/实缴资本', '绿色金融指数', '单位GDP能耗', '碳排放量/碳汇',
    '新增新质生产力用地面积', '绿色专利授权数量', '专精特新小巨人企业专利应用',
    '科研机构增长率', '产业结构高级化指数', '新增新质生产力相关企业数量',
    '高新技术企业占比', '居民收入or增速or收入支出比', '城乡收入差距',
    '建成区绿地覆盖率', '环境处罚', '通过应用绿色技术的企业比例'
]

# 定义负向指标
negative_indicators = ['单位GDP能耗', '碳排放量/碳汇', '环境处罚', '城乡收入差距']

# 检查选择的列是否都存在于数据框中
missing_columns = [col for col in columns_to_normalize if col not in df.columns]
if missing_columns:
    print(f"警告：以下列不存在于数据框中：{', '.join(missing_columns)}")
    columns_to_normalize = [col for col in columns_to_normalize if col in df.columns]

# 4. 处理无穷大的值
for col in columns_to_normalize:
    df[col] = df[col].replace([np.inf, -np.inf], np.nan)
    df[col] = df[col].fillna(df[col].mean())
    # 处理过大的值（例如，将超过10倍标准差的值设置为10倍标准差）
    mean = df[col].mean()
    std = df[col].std()
    df[col] = df[col].clip(lower=mean - 10 * std, upper=mean + 10 * std)

# 5. 进行归一化处理
scaler = MinMaxScaler()
df_normalized = pd.DataFrame()

for col in columns_to_normalize:
    if col in negative_indicators:
        # 对负向指标进行反向处理
        df_normalized[col] = 1 - scaler.fit_transform(df[[col]]).flatten()
    else:
        # 对正向指标进行常规归一化
        df_normalized[col] = scaler.fit_transform(df[[col]]).flatten()

# 将归一化后的数据合并回原始数据框
df[columns_to_normalize] = df_normalized
# 去除列名中有.1的列
columns_to_remove = [col for col in df.columns if '.1' in col]
df = df.drop(columns=columns_to_remove)

print(f"已删除以下列：{', '.join(columns_to_remove)}")

# 6. 保存归一化后的数据
output_file_path = r"J:\BaiduNetdiskDownload\05马克\8亿+企业查数据\03指标计算\企业查区县统计合并_新指标和原始指标merged_data归一化.xlsx"
df.to_excel(output_file_path, index=False)

# 7. 为每一列添加蓝色渐变条件格式
wb = load_workbook(output_file_path)
ws = wb.active

for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
    column_letter = col[0].column_letter
    if col[0].value in columns_to_normalize:
        ws.conditional_formatting.add(
            f'{column_letter}2:{column_letter}{ws.max_row}',
            ColorScaleRule(start_type='min', start_color='DEEBF7',
                           mid_type='percentile', mid_value=50, mid_color='9ECAE1',
                           end_type='max', end_color='3182BD')
        )

# 保存带有条件格式的Excel文件
wb.save(output_file_path)

print("归一化处理完成，并已保存带有条件格式的Excel文件至:", output_file_path)

# 8. 检查归一化结果
print("\n归一化后的数据统计信息：")
print(df[columns_to_normalize].describe())

# 9. 检查是否有缺失值
missing_values = df[columns_to_normalize].isnull().sum()
if missing_values.sum() > 0:
    print("\n警告：归一化后存在缺失值：")
    print(missing_values[missing_values > 0])
else:
    print("\n归一化后没有缺失值。")
